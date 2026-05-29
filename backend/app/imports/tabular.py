"""
Tabular helpers shared by the CSV/XLSX importers.

Vendor exports are messy: header names drift between product versions, dates
come in three formats, speeds in m/s or km/h. These helpers absorb that so the
per-vendor importers stay declarative (just an alias map per field).
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple


def _norm_key(s: str) -> str:
    out = re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower())
    return out.strip("_")


def read_rows(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return (original_headers, rows). Rows are dicts keyed by *normalised* header.

    Supports .csv/.tsv (sniffed delimiter) and .xlsx (first sheet, openpyxl).
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    return _read_csv(path)


def _read_csv(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    raw = path.read_bytes()
    # Tolerate BOM and latin-1 fallback (Polar/older exports).
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover
        text = raw.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows_raw = [r for r in reader if any(c.strip() for c in r)]
    if not rows_raw:
        return [], []
    headers = rows_raw[0]
    norm = [_norm_key(h) for h in headers]
    out: List[Dict[str, Any]] = []
    for r in rows_raw[1:]:
        out.append({norm[i]: (r[i] if i < len(r) else None) for i in range(len(norm))})
    return headers, out


def _read_xlsx(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("Falta openpyxl para leer .xlsx (pip install openpyxl)") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h) if h is not None else "" for h in header_row]
    norm = [_norm_key(h) for h in headers]
    out: List[Dict[str, Any]] = []
    for r in rows_iter:
        if r is None or all(c is None for c in r):
            continue
        out.append({norm[i]: (r[i] if i < len(r) else None) for i in range(len(norm))})
    wb.close()
    return headers, out


def pick(row: Dict[str, Any], *aliases: str) -> Optional[Any]:
    """First non-empty value among the given (human) header aliases."""
    for a in aliases:
        key = _norm_key(a)
        if key in row:
            v = row[key]
            if v is not None and str(v).strip() != "":
                return v
    return None


def to_float(v: Any) -> Optional[float]:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def to_int(v: Any) -> Optional[int]:
    f = to_float(v)
    return int(round(f)) if f is not None else None


def mps_to_kmh(v: Optional[float]) -> Optional[float]:
    return round(v * 3.6, 2) if v is not None else None


def speed_to_kmh(v: Any, header_hint: str = "") -> Optional[float]:
    """Normalise a max-speed value to km/h.

    Uses the header text as a unit hint; if it says m/s (or the magnitude is
    implausibly low for km/h, < 15), assume m/s and convert.
    """
    f = to_float(v)
    if f is None:
        return None
    hint = header_hint.lower()
    if "m/s" in hint or "mps" in hint:
        return mps_to_kmh(f)
    if "km/h" in hint or "kmh" in hint or "kph" in hint:
        return round(f, 2)
    # No unit hint: a footballer's max is ~25-37 km/h or ~7-10 m/s.
    return mps_to_kmh(f) if f < 15 else round(f, 2)


def parse_date(v: Any) -> Optional[date]:
    """Parse a date/datetime from common vendor formats and Excel serials."""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Excel serial date (days since 1899-12-30)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
        except (ValueError, OverflowError):
            return None

    s = str(v).strip()
    # Strip a trailing time component if present.
    s_date = s.split("T")[0].split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s_date, fmt).date()
        except ValueError:
            continue
    # Last resort: dateutil (ships with pandas), day-first for non-US vendors.
    try:
        from dateutil import parser as dtp
        return dtp.parse(s, dayfirst=True).date()
    except (ValueError, ImportError, OverflowError):
        return None


def header_for(headers: Sequence[str], *aliases: str) -> str:
    """Return the original header text matching one of the aliases (for unit hints)."""
    norm_aliases = {_norm_key(a) for a in aliases}
    for h in headers:
        if _norm_key(h) in norm_aliases:
            return h
    return ""
