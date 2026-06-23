from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
import base64
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from ....core.database import get_db
from ....core.deps import get_current_user, require_roles, get_current_club_id, scoped_query
from ....models.player import Player, PlayerStatus
from ....models.category import Category
from ....models.user import UserRole, User
from ....schemas.player import PlayerCreate, PlayerUpdate, PlayerOut, PlayerSummary

router = APIRouter()


@router.get("/", response_model=List[PlayerSummary])
def list_players(
    category_id: Optional[int] = None,
    status: Optional[PlayerStatus] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Player).filter(Player.is_active == True)
    q = scoped_query(q, Player, current_user)
    if category_id:
        q = q.filter(Player.category_id == category_id)
    if status:
        q = q.filter(Player.status == status)
    if search:
        q = q.filter(
            (Player.first_name.ilike(f"%{search}%")) | (Player.last_name.ilike(f"%{search}%"))
        )
    return q.order_by(Player.last_name).offset(skip).limit(limit).all()


@router.get("/{player_id}", response_model=PlayerOut)
def get_player(player_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Player).options(joinedload(Player.category)).filter(Player.id == player_id)
    q = scoped_query(q, Player, current_user)
    player = q.first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")
    return player


@router.post("/", response_model=PlayerOut, status_code=status.HTTP_201_CREATED)
def create_player(
    data: PlayerCreate,
    db: Session = Depends(get_db),
    club_id: int = Depends(get_current_club_id),
    _=Depends(require_roles(UserRole.ADMIN, UserRole.COACH)),
):
    # Verify category belongs to current club
    cat = db.query(Category).filter(Category.id == data.category_id, Category.club_id == club_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Categoría inválida para este club")
    payload = data.model_dump()
    payload["club_id"] = club_id
    player = Player(**payload)
    db.add(player)
    db.commit()
    db.refresh(player)
    return db.query(Player).options(joinedload(Player.category)).filter(Player.id == player.id).first()


@router.patch("/{player_id}", response_model=PlayerOut)
def update_player(
    player_id: int,
    data: PlayerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_roles(UserRole.ADMIN, UserRole.COACH, UserRole.KINESIOLOGIST)),
):
    q = db.query(Player).filter(Player.id == player_id)
    q = scoped_query(q, Player, current_user)
    player = q.first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")
    for field, value in data.model_dump(exclude_none=True).items():
        # Block tenant escalation
        if field == "club_id":
            continue
        setattr(player, field, value)
    db.commit()
    db.refresh(player)
    return db.query(Player).options(joinedload(Player.category)).filter(Player.id == player_id).first()


@router.post("/{player_id}/photo", response_model=PlayerOut)
def upload_player_photo(
    player_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_roles(UserRole.ADMIN, UserRole.COACH)),
):
    q = db.query(Player).filter(Player.id == player_id)
    q = scoped_query(q, Player, current_user)
    player = q.first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")
    if file.content_type not in ("image/jpeg", "image/png", "image/webp"):
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes JPEG, PNG o WebP")
    content = file.file.read()
    if len(content) > 1_000_000:  # 1 MB max
        raise HTTPException(status_code=400, detail="La imagen no puede superar 1 MB")
    b64 = base64.b64encode(content).decode("utf-8")
    player.photo_url = f"data:{file.content_type};base64,{b64}"
    db.commit()
    db.refresh(player)
    return db.query(Player).options(joinedload(Player.category)).filter(Player.id == player_id).first()


@router.delete("/{player_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_player(
    player_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_roles(UserRole.ADMIN)),
):
    q = db.query(Player).filter(Player.id == player_id)
    q = scoped_query(q, Player, current_user)
    player = q.first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")
    player.is_active = False
    db.commit()


@router.post("/import", status_code=status.HTTP_200_OK)
def import_players(
    file: UploadFile = File(...),
    category_id: int = Form(...),
    db: Session = Depends(get_db),
    club_id: int = Depends(get_current_club_id),
    current_user: User = Depends(get_current_user),
    _=Depends(require_roles(UserRole.ADMIN, UserRole.COACH)),
):
    import io
    import csv
    import re
    from datetime import date, datetime
    from ....models.player import PlayerPosition, DominantFoot

    def _normalize_key(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower()).strip("_")

    # Verify category belongs to current club
    cat = db.query(Category).filter(Category.id == category_id, Category.club_id == club_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Categoría inválida para este club")

    content = file.file.read()
    filename = file.filename or ""
    rows = []

    try:
        if filename.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                raise HTTPException(status_code=400, detail="Archivo Excel vacío")
            headers = [str(h) if h is not None else "" for h in header_row]
            norm_headers = [_normalize_key(h) for h in headers]
            for r in rows_iter:
                if r is None or all(c is None for c in r):
                    continue
                rows.append({norm_headers[i]: (r[i] if i < len(r) else None) for i in range(len(norm_headers))})
            wb.close()
        else:
            # CSV / TSV fallback
            text = ""
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    text = content.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if not text:
                text = content.decode("utf-8", errors="replace")
            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(io.StringIO(text), dialect)
            rows_raw = [r for r in reader if any(c.strip() for c in r)]
            if not rows_raw:
                raise HTTPException(status_code=400, detail="Archivo CSV vacío")
            headers = rows_raw[0]
            norm_headers = [_normalize_key(h) for h in headers]
            for r in rows_raw[1:]:
                rows.append({norm_headers[i]: (r[i] if i < len(r) else None) for i in range(len(norm_headers))})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al analizar archivo: {str(e)}")

    def get_field(r: dict, aliases: list[str]) -> any:
        for a in aliases:
            k = _normalize_key(a)
            if k in r:
                v = r[k]
                if v is not None and str(v).strip() != "":
                    return v
        return None

    imported = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows):
        first_name = get_field(row, ["first_name", "nombre", "nombres"])
        last_name = get_field(row, ["last_name", "apellido", "apellidos"])

        if not first_name or not last_name:
            full_name = get_field(row, ["full_name", "nombre_completo", "nombre_y_apellido"])
            if full_name:
                parts = str(full_name).strip().split(" ", 1)
                first_name = parts[0]
                last_name = parts[1] if len(parts) > 1 else "S/A"
            else:
                skipped += 1
                errors.append(f"Fila {idx+2}: Faltan nombres o apellidos")
                continue

        # Date of Birth
        dob_raw = get_field(row, ["date_of_birth", "fecha_nacimiento", "nacimiento", "dob"])
        dob = None
        if dob_raw:
            from datetime import timedelta
            if isinstance(dob_raw, (datetime, date)):
                dob = dob_raw
            elif isinstance(dob_raw, (int, float)):
                try:
                    dob = (datetime(1899, 12, 30) + timedelta(days=float(dob_raw))).date()
                except Exception:
                    pass
            else:
                s_date = str(dob_raw).strip().split("T")[0].split(" ")[0]
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"):
                    try:
                        dob = datetime.strptime(s_date, fmt).date()
                        break
                    except ValueError:
                        continue
        if not dob:
            dob = date(2000, 1, 1)

        # Position
        pos_raw = get_field(row, ["position", "posicion"])
        pos = PlayerPosition.CENTRAL_MID
        if pos_raw:
            pos_str = _normalize_key(str(pos_raw))
            eng_map = {
                "goalkeeper": PlayerPosition.GOALKEEPER, "gk": PlayerPosition.GOALKEEPER, "arquero": PlayerPosition.GOALKEEPER, "portero": PlayerPosition.GOALKEEPER,
                "center_back": PlayerPosition.CENTER_BACK, "cb": PlayerPosition.CENTER_BACK, "defensa_central": PlayerPosition.CENTER_BACK, "central": PlayerPosition.CENTER_BACK,
                "left_back": PlayerPosition.LEFT_BACK, "lb": PlayerPosition.LEFT_BACK, "lateral_izquierdo": PlayerPosition.LEFT_BACK,
                "right_back": PlayerPosition.RIGHT_BACK, "rb": PlayerPosition.RIGHT_BACK, "lateral_derecho": PlayerPosition.RIGHT_BACK,
                "defensive_mid": PlayerPosition.DEFENSIVE_MID, "dm": PlayerPosition.DEFENSIVE_MID, "volante_defensivo": PlayerPosition.DEFENSIVE_MID, "contencion": PlayerPosition.DEFENSIVE_MID,
                "central_mid": PlayerPosition.CENTRAL_MID, "cm": PlayerPosition.CENTRAL_MID, "volante_mixto": PlayerPosition.CENTRAL_MID, "mediocampista": PlayerPosition.CENTRAL_MID,
                "attacking_mid": PlayerPosition.ATTACKING_MID, "am": PlayerPosition.ATTACKING_MID, "volante_creacion": PlayerPosition.ATTACKING_MID, "enganche": PlayerPosition.ATTACKING_MID,
                "left_wing": PlayerPosition.LEFT_WING, "lw": PlayerPosition.LEFT_WING, "puntero_izquierdo": PlayerPosition.LEFT_WING, "extremo_izquierdo": PlayerPosition.LEFT_WING,
                "right_wing": PlayerPosition.RIGHT_WING, "rw": PlayerPosition.RIGHT_WING, "puntero_derecho": PlayerPosition.RIGHT_WING, "extremo_derecho": PlayerPosition.RIGHT_WING,
                "center_forward": PlayerPosition.CENTER_FORWARD, "cf": PlayerPosition.CENTER_FORWARD, "delantero_centro": PlayerPosition.CENTER_FORWARD, "delantero": PlayerPosition.CENTER_FORWARD, "nueve": PlayerPosition.CENTER_FORWARD,
            }
            if pos_str in eng_map:
                pos = eng_map[pos_str]

        # Dominant Foot
        foot_raw = get_field(row, ["dominant_foot", "pie", "pie_dominante", "perfil"])
        foot = DominantFoot.RIGHT
        if foot_raw:
            foot_str = _normalize_key(str(foot_raw))
            if foot_str in ("left", "izquierdo", "zurdo", "i", "izq"):
                foot = DominantFoot.LEFT
            elif foot_str in ("both", "ambos", "ambidiestro", "a"):
                foot = DominantFoot.BOTH

        jersey = get_field(row, ["jersey_number", "numero", "dorsal", "camiseta"])
        jersey_num = None
        if jersey is not None:
            try:
                jersey_num = int(float(str(jersey).strip()))
            except Exception:
                pass

        height = get_field(row, ["height_cm", "height", "altura", "estatura"])
        height_cm = None
        if height is not None:
            try:
                height_cm = float(str(height).replace(",", ".").strip())
                if height_cm < 3.0:
                    height_cm = height_cm * 100.0
            except Exception:
                pass

        weight = get_field(row, ["weight_kg", "weight", "peso"])
        weight_kg = None
        if weight is not None:
            try:
                weight_kg = float(str(weight).replace(",", ".").strip())
            except Exception:
                pass

        document_id = get_field(row, ["document_id", "rut", "run", "dni", "cedula"])
        email = get_field(row, ["email", "correo", "mail"])
        phone = get_field(row, ["phone", "telefono", "celular"])

        player = Player(
            first_name=str(first_name).strip(),
            last_name=str(last_name).strip(),
            date_of_birth=dob,
            position=pos,
            dominant_foot=foot,
            jersey_number=jersey_num,
            height_cm=height_cm,
            weight_kg=weight_kg,
            document_id=str(document_id).strip() if document_id else None,
            email=str(email).strip() if email else None,
            phone=str(phone).strip() if phone else None,
            club_id=club_id,
            category_id=category_id,
            status=PlayerStatus.AVAILABLE,
        )
        db.add(player)
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}
